from openpyxl import load_workbook
from collections import Counter

big_spisok = []  # ............................. список для хранения всех(!) значений из таблицы
results_of_numbers = {}  # .......................словарь для хранения обработанных данных из таблицы
wb = load_workbook('--file name---.xlsx')
ws = wb['Worksheet']
number_of_stroka = 0  # ......................... номер строки при выводе на консоль
exel_stroka_number = 704  # ....................... номер строки(с какой мы начнем брать данные) в файле excel
lenght_of_table = 1335  # ...................... до какой строки нам нужно брать данные
start_position = exel_stroka_number  # ............................старт

while exel_stroka_number <= lenght_of_table:
    exel_stroka_number += 1
    number_of_stroka += 1
    x = f'C{exel_stroka_number}'
    y = f'H{exel_stroka_number}'
    for row in ws[x:y]:
        string = ''
        for cell in row:
            string += str(cell.value) + ' '
            string2 = string.split()
    string2 = list(map(int, string2))
    string2.sort()
    big_spisok.extend(string2)
    results_of_numbers[f'List{number_of_stroka}'] = string2  # ...внесли все данные в словарь
    # print(f'List#{number_of_stroka}: {string2}') # (распечатка данных из таблицы)

big_spisok.sort()
c_big_spisok = Counter(big_spisok)  # ........... сортировка по количеству совадающих элементов
# print(c_big_spisok)
# print(results_of_numbers)

